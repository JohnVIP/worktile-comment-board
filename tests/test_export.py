#!/usr/bin/env python3
"""一键导出 Excel 的单元测试（mock 掉 Worktile 网络层，不依赖真实凭证）。

覆盖：
1. _collect_tasks_for_export 的 owner + keyword 过滤
2. fetch_all_comments 的并发聚合与单条失败兜底
3. build_workbook 生成 2 个 sheet（含评论）与 1 个 sheet（仅任务），字段正确
4. /api/export 路由：board 视图、含评论 / 仅任务、响应为合法 xlsx 且带下载头
"""

import io
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

# 让测试能 import 项目根目录的 wt_client / app / exporter
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from wt_client import WorktileClient  # noqa: E402
from exporter import build_workbook  # noqa: E402
import app as app_module  # noqa: E402


# --------------------------------------------------------------------------- 假数据
def _make_task(tid, title, uid, project="P1", due=None, done=False):
    return {
        "task_id": tid, "identifier": f"T-{tid}", "title": title,
        "project_id": project, "project_name": f"项目{project}",
        "assignee": uid or "未分配", "is_completed": done,
        "updated_at_str": "2026-08-20 10:00", "due_at_str": "-",
        "due_at": due, "overdue_days": None, "_assignee_uid": uid,
    }


TASKS = [
    _make_task("1", "写方案", "u1"),
    _make_task("2", "评审方案", "u2"),
    _make_task("3", "未分配任务", None),
    _make_task("4", "写方案第二版", "u1"),
]


class FakeClient(WorktileClient):
    """用内存假数据替换网络层；不发起任何请求。"""

    def __init__(self):
        # 不调用父类 __init__（父类会要求凭证）；手动塞必要属性
        self._FETCH_PAGE = 200
        self._file_cache = {}

    def _iter_all_tasks(self, projects, project_id, page_size, member_map,
                        force=False):
        for t in TASKS:
            yield t, t["project_name"]

    def get_task_comments(self, task_id, member_map=None):
        return [
            {"author": "小王", "content": f"{task_id} 的评论",
             "created_at_str": "2026-08-20 09:00", "attachments": []},
        ]


# --------------------------------------------------------------------------- 1. 收集过滤
def test_collect_tasks_filter():
    c = FakeClient()
    # 不过滤
    all_t = c._collect_tasks_for_export([{"id": "P1"}], "P1", None, "", {})
    assert len(all_t) == 4
    # owner 过滤（uid）
    u1 = c._collect_tasks_for_export([{"id": "P1"}], "P1", "u1", "", {})
    assert {t["task_id"] for t in u1} == {"1", "4"}
    # 未分配
    un = c._collect_tasks_for_export([{"id": "P1"}], "P1", "__unassigned__", "", {})
    assert {t["task_id"] for t in un} == {"3"}
    # keyword（OR，标题子串）
    kw = c._collect_tasks_for_export([{"id": "P1"}], "P1", None, "方案", {})
    assert {t["task_id"] for t in kw} == {"1", "2", "4"}
    # owner + keyword 组合
    combo = c._collect_tasks_for_export([{"id": "P1"}], "P1", "u1", "第二版", {})
    assert {t["task_id"] for t in combo} == {"4"}


# --------------------------------------------------------------------------- 2. 评论聚合
def test_fetch_all_comments():
    c = FakeClient()
    res = c.fetch_all_comments(["1", "2"], {})
    assert set(res.keys()) == {"1", "2"}
    assert res["1"][0]["author"] == "小王"

    # 单条失败兜底：让 task_id="boom" 抛错
    def _boom(tid, member_map=None):
        if tid == "boom":
            raise RuntimeError("模拟限流")
        return [{"author": "x", "content": "ok", "created_at_str": "-", "attachments": []}]
    c.get_task_comments = _boom
    res2 = c.fetch_all_comments(["a", "boom"], {})
    assert res2["a"][0]["content"] == "ok"
    assert res2["boom"] == []  # 失败降级为空列表，不影响其它


# --------------------------------------------------------------------------- 3. workbook
def test_build_workbook_with_comments():
    task_rows = [
        {"task_id": "1", "identifier": "T-1", "title": "写方案", "project_name": "项目P1",
         "assignee": "u1", "status": "进行中", "is_completed": False,
         "desc": "方案描述", "desc_images": ["https://img/a.png", "https://img/b.png"],
         "updated_at_str": "2026-08-20 10:00", "start_at_str": "2026-08-19 09:00",
         "due_at_str": "-", "overdue_days": None, "comment_count": 1},
    ]
    comment_rows = [
        {"task_id": "1", "title": "写方案", "project_name": "项目P1", "author": "小王",
         "created_at_str": "2026-08-20 09:00", "content": "评论内容", "attachments": []},
    ]
    wb_bytes, sheets = build_workbook(task_rows, comment_rows, with_comments=True)
    assert sheets == 2
    wb = load_workbook(io.BytesIO(wb_bytes))
    assert wb.sheetnames == ["任务", "评论"]
    ws = wb["任务"]
    assert [c.value for c in ws[1]] == ["任务ID", "编号", "标题", "描述", "描述图片", "项目",
                                        "负责人", "状态", "是否完成", "更新时间", "开始时间",
                                        "截止时间", "逾期天数", "评论数"]
    assert ws.cell(row=2, column=1).value == "1"
    assert ws.cell(row=2, column=4).value == "方案描述"   # 描述
    assert ws.cell(row=2, column=5).value == "https://img/a.png\nhttps://img/b.png"  # 描述图片
    assert ws.cell(row=2, column=8).value == "进行中"     # 状态
    assert ws.cell(row=2, column=11).value == "2026-08-19 09:00"  # 开始时间
    assert ws.cell(row=2, column=13).value == "-"         # 逾期天数（描述/状态/开始时间 3 列插入后顺延）
    cs = wb["评论"]
    assert cs.cell(row=2, column=4).value == "小王"
    assert cs.cell(row=2, column=6).value == "评论内容"


def test_build_workbook_tasks_only():
    task_rows = [{"task_id": "1", "title": "t", "comment_count": 0}]
    wb_bytes, sheets = build_workbook(task_rows, [], with_comments=False)
    assert sheets == 1
    wb = load_workbook(io.BytesIO(wb_bytes))
    assert wb.sheetnames == ["任务"]


# --------------------------------------------------------------------------- 4. 路由
def test_export_route_board():
    fake_client = FakeClient()
    fake_session = {
        "projects": [{"id": "P1", "name": "项目P1"}],
        "client": fake_client,
        "member_map": {},
    }
    app_module._require_session = lambda: fake_session
    client = app_module.app.test_client()

    # 仅任务
    r = client.get("/api/export?view=board&project_id=P1&with_comments=0")
    assert r.status_code == 200
    assert r.headers["Content-Disposition"].startswith('attachment; filename="worktile_board_')
    assert r.data[:2] == b"PK"  # xlsx(zip) 魔数
    assert int(r.headers["Content-Length"]) > 0

    # 含评论
    r2 = client.get("/api/export?view=board&project_id=P1&with_comments=1")
    assert r2.status_code == 200
    wb = load_workbook(io.BytesIO(r2.data))
    assert wb.sheetnames == ["任务", "评论"]
    # 任务 sheet 应含全部 4 条（无 owner/keyword 过滤）
    assert wb["任务"].max_row == 5  # 1 表头 + 4 数据


def test_export_route_overdue_owner_filter():
    """overdue 视图：复用 _compute_overdue，按 owner 过滤；badge/total 一致。"""
    fake_client = FakeClient()
    fake_session = {
        "projects": [{"id": "P1", "name": "项目P1"}],
        "client": fake_client,
        "member_map": {},
    }
    app_module._require_session = lambda: fake_session
    client = app_module.app.test_client()
    r = client.get("/api/export?view=overdue&project_id=P1&owner=u1&with_comments=0")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.data))
    # 不带 due → _compute_overdue 全部排除（due=None），任务 sheet 仅表头
    assert wb["任务"].max_row == 1


# --------------------------------------------------------------------------- 5. 异步作业流（进度轮询 + 下载）
def test_export_job_flow():
    """start → progress(轮询至 done) → download，校验进度计数与最终 xlsx。"""
    import time as _t  # noqa: E402

    fake_client = FakeClient()
    fake_session = {
        "projects": [{"id": "P1", "name": "项目P1"}],
        "client": fake_client,
        "member_map": {},
    }
    app_module._require_session = lambda: fake_session
    client = app_module.app.test_client()

    # 启动异步作业（board + 含评论）
    r = client.post("/api/export/start",
                    json={"view": "board", "project_id": "P1", "with_comments": True})
    assert r.status_code == 200
    data = r.get_json()
    assert data["ok"] and data["job_id"]
    job_id = data["job_id"]

    # 轮询直到完成（后台线程，固定假数据很快）
    final = None
    for _ in range(50):
        p = client.get(f"/api/export/progress?job_id={job_id}").get_json()
        if p["status"] in ("done", "error"):
            final = p
            break
        _t.sleep(0.1)
    assert final is not None, "导出作业未在预期时间内完成"
    assert final["status"] == "done", final.get("error")
    assert final["task_count"] == 4
    assert final["comment_count"] == 4  # 每个任务 1 条评论，进度回调应累计到 4

    # 下载文件并校验为合法 xlsx（任务 + 评论 两个 sheet）
    dl = client.get(f"/api/export/download?job_id={job_id}")
    assert dl.status_code == 200
    assert dl.data[:2] == b"PK"
    wb = load_workbook(io.BytesIO(dl.data))
    assert wb.sheetnames == ["任务", "评论"]

    # 下载后作业应被清理：progress 查询返回 404
    p2 = client.get(f"/api/export/progress?job_id={job_id}")
    assert p2.status_code == 404


# --------------------------------------------------------------------------- 6. 逾期天数计算
def test_compute_overdue_days():
    now = 1_000_000_000
    # 已过期 3 天且未完成 → 3
    assert app_module.compute_overdue_days(now - 3 * 86400, False, now) == 3
    # 刚好到期（due == now）→ None（不算逾期）
    assert app_module.compute_overdue_days(now, False, now) is None
    # 未到期（未来）→ None
    assert app_module.compute_overdue_days(now + 86400, False, now) is None
    # 已完成 → None（无论是否过期）
    assert app_module.compute_overdue_days(now - 10 * 86400, True, now) is None
    # 无截止时间 → None
    assert app_module.compute_overdue_days(None, False, now) is None


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-v"]))

#!/usr/bin/env python3
"""把任务 + 评论数据组装成 .xlsx 工作簿（openpyxl）。

设计要点：
- 任务 sheet 永远存在；评论 sheet 仅在 with_comments=True 时创建。
- 字段全部来自现有 normalize_task / get_task_comments 的输出，无需新的解析逻辑。
- 表头加粗白字蓝底；评论正文列自动换行；冻结首行；列宽按内容自适应（封顶 60）。
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_WRAP_TOP = Alignment(vertical="top", wrap_text=True)


def _attachment_names(attachments):
    """把评论 attachments 列表里的文件名拼成 ; 分隔字符串。"""
    if not attachments:
        return ""
    names = []
    for a in attachments:
        if isinstance(a, dict):
            n = a.get("name") or a.get("title") or ""
            if n:
                names.append(str(n))
    return "; ".join(names)


def _autosize(ws, ncols, max_width=60):
    """按每列最长内容设置列宽（中文按宽度粗略计算，封顶 max_width）。"""
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        longest = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            # 中文/全角字符按 2 宽度估算
            w = sum(2 if ord(ch) > 0x2E80 else 1 for ch in s)
            if w > longest:
                longest = w
        ws.column_dimensions[letter].width = min(max(longest + 2, 8), max_width)


def build_workbook(task_rows, comment_rows, with_comments=True):
    """生成 xlsx 字节流，返回 (bytes, sheet_count)。

    task_rows 每项 dict（键来源 normalize_task + 导出聚合）：
        task_id, identifier, title, project_name, assignee, is_completed,
        updated_at_str, due_at_str, overdue_days, comment_count
    comment_rows 每项 dict：
        task_id, title, project_name, author, created_at_str, content, attachments(list)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "任务"

    task_headers = ["任务ID", "编号", "标题", "描述", "描述图片", "项目", "负责人", "状态",
                    "是否完成", "更新时间", "开始时间", "截止时间", "逾期天数", "评论数"]
    ws.append(task_headers)
    for c in range(1, len(task_headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    for t in task_rows:
        overdue = t.get("overdue_days")
        overdue_str = "-" if overdue is None else str(overdue)
        # 描述图片：URL 列表按换行拼成多行，便于在 Excel 内逐个查看
        desc_imgs = t.get("desc_images") or []
        desc_img_str = "\n".join(str(u) for u in desc_imgs) if desc_imgs else ""
        ws.append([
            t.get("task_id", ""),
            t.get("identifier", ""),
            t.get("title", ""),
            t.get("desc", ""),
            desc_img_str,
            t.get("project_name", ""),
            t.get("assignee", ""),
            t.get("status", ""),
            "是" if t.get("is_completed") else "否",
            t.get("updated_at_str") or "-",
            t.get("start_at_str") or "-",
            t.get("due_at_str") or "-",
            overdue_str,
            t.get("comment_count", 0),
        ])
    # 描述（第 4 列）与描述图片（第 5 列）自动换行，便于阅读长文本 / 多链接
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=4).alignment = _WRAP_TOP
        ws.cell(row=r, column=5).alignment = _WRAP_TOP
    _autosize(ws, len(task_headers))
    ws.freeze_panes = "A2"

    sheet_count = 1

    if with_comments:
        cs = wb.create_sheet("评论")
        comment_headers = ["任务ID", "任务标题", "项目", "评论人",
                           "评论时间", "评论内容", "附件"]
        cs.append(comment_headers)
        for c in range(1, len(comment_headers) + 1):
            cell = cs.cell(row=1, column=c)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(vertical="center")
        for cm in comment_rows:
            cs.append([
                cm.get("task_id", ""),
                cm.get("title", ""),
                cm.get("project_name", ""),
                cm.get("author", ""),
                cm.get("created_at_str") or "-",
                cm.get("content", ""),
                _attachment_names(cm.get("attachments")),
            ])
        # 评论内容列（第 6 列）换行，便于阅读长评论
        for r in range(2, cs.max_row + 1):
            cs.cell(row=r, column=6).alignment = _WRAP_TOP
        _autosize(cs, len(comment_headers))
        cs.freeze_panes = "A2"
        sheet_count = 2

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), sheet_count
